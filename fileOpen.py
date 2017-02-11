from openpyxl import load_workbook
from tkinter.filedialog import askopenfilename

def PK():
    fileName = askopenfilename( filetypes = (("CSV files", "*.csv"),("All files", "*.*")))
    f = open(fileName, 'r')
    x=f.read()
    f.close()
    fileList = x.split('\n')
    a = fileList[0].split(',')
    nutrients = len(a) - 1
    if nutrients == 3:
        d,d1 = [],[]
        m,m1 = [],[]
        y,y1 = [],[]
        ObsP,ObsK,ObsS,MassP, MassK, MassS=[],[],[],[],[],[]
        k = 0
        for eachLine in fileList:
            date,obsP,obsK,obsS = eachLine.split(',')
            if len(date) >= 1 and k < 0.5:
                t0,t1,t2=date.split('/')
                d.append(float(t0))
                m.append(float(t1))
                y.append(float(t2))
                ObsP.append(float(obsP))
                ObsK.append(float(obsK))
                ObsS.append(float(obsS))
            elif len(date) < 1 and k < 0.5:
                k = k + 1
            elif len(date) >= 1 and k > 0.5:
                t0,t1,t2=date.split('/')
                d1.append(float(t0))
                m1.append(float(t1))
                y1.append(float(t2))
                MassP.append(float(obsP))
                MassK.append(float(obsK))
                MassS.append(float(obsS))
        TObs = [xx-1+30*(yy-1)+365*zz for xx,yy,zz in zip (d, m, y)]
        Tfert = [xx-1+30*(yy-1)+365*zz for xx,yy,zz in zip (d1, m1, y1)]
        return TObs,ObsP,ObsK,ObsS, Tfert, MassP,MassK,MassS
    else:
        return 'No Data'



def PKXl():
    fileName = askopenfilename( filetypes = [("Excel files", "*.xlsx"),("All files", "*.*")])
    f = load_workbook(fileName)
    sheet=f.active
    Tobs, ObsP, ObsK, ObsS, Tfert, MassP, MassK, MassS = [],[],[],[],[],[],[],[]
    TotalR = sheet.max_row
    for i in range(2,TotalR+1):
        try:
            Tobs.append(sheet.cell(row=i, column=1).value.day+30*(sheet.cell(row=i, column=1).value.month-1)+sheet.cell(row=i, column=1).value.year*365)
        except(AttributeError):
            pass
        try:
            ObsP.append(sheet.cell(row=i, column=2).value)
        except(AttributeError):
            pass
        try:
            ObsK.append(sheet.cell(row=i, column=3).value)
        except(AttributeError):
            pass
        try:
            ObsS.append(sheet.cell(row=i, column=4).value)
        except(AttributeError):
            pass
        try:
            Tfert.append(sheet.cell(row=i, column=5).value.day+30*(sheet.cell(row=i, column=5).value.month-1)+sheet.cell(row=i, column=5).value.year*365)
        except(AttributeError):
            pass
        try:
            MassP.append(sheet.cell(row=i, column=6).value)
        except(AttributeError):
            pass
        try:
            MassK.append(sheet.cell(row=i, column=7).value)
        except(AttributeError):
            pass
        try:
            MassS.append(sheet.cell(row=i, column=8).value)
        except(AttributeError):
            pass
    
    return Tobs,ObsP,ObsK,ObsS, Tfert, MassP,MassK,MassS
