from openpyxl import load_workbook
import numpy as np
from tkinter.filedialog import askopenfilename

def PK():
    fileName = askopenfilename( filetypes = (("CSV files", "*.csv"),("All files", "*.*")))
    f = open(fileName, 'r')
    x=f.read()
    f.close()
    fileList = x.split('\n')
    a = fileList[0].split(',')
    trials = len(a) - 1
    if trials == 1:
        d,d1 = [],[]
        m,m1 = [],[]
        y,y1 = [],[]
        Obs,mass=[],[]
        k = 0
        for eachLine in fileList:
            date,obs = eachLine.split(',')
            if len(date) >= 1 and k < 0.5:
                t0,t1,t2 = date.split('/')
                d.append(float(t0))
                m.append(float(t1))
                y.append(float(t2))
                Obs.append(float(obs))
            elif len(date) < 1 and k < 0.5:
                k = k + 1
            elif len(date) >= 1 and k > 0.5:
                t0,t1,t2 = date.split('/')
                d1.append(float(t0))
                m1.append(float(t1))
                y1.append(float(t2))
                mass.append(float(obs))
        TObs = [xx-1+30*(yy-1)+365*zz for xx,yy,zz in zip (d, m, y)]
        Tfert = [xx-1+30*(yy-1)+365*zz for xx,yy,zz in zip (d1, m1, y1)]
        return np.array(TObs),np.array(Obs), np.array(Tfert), np.array(mass)
    elif trials == 2:
        d,d1 = [],[]
        m,m1 = [],[]
        y,y1 = [],[]
        Obs1,mass1 = [],[]
        Obs2,mass2 = [],[]
        k = 0
        for eachLine in fileList:
            date,obs1,obs2 = eachLine.split(',')
            if len(date) >= 1 and k < 0.5:
                t0,t1,t2=date.split('/')
                d.append(float(t0))
                m.append(float(t1))
                y.append(float(t2))
                Obs1.append(float(obs1))
                Obs2.append(float(obs2))
            elif len(date) < 1 and k < 0.5:
                k = k + 1
            elif len(date) >= 1 and k > 0.5:
                t0,t1,t2=date.split('/')
                d1.append(float(t0))
                m1.append(float(t1))
                y1.append(float(t2))
                mass1.append(float(obs1))
                mass2.append(float(obs2))
        TObs = [xx-1+30*(yy-1)+365*zz for xx,yy,zz in zip (d, m, y)]
        Tfert = [xx-1+30*(yy-1)+365*zz for xx,yy,zz in zip (d1, m1, y1)]
        return np.array(TObs), np.column_stack((Obs1,Obs2)),np.array(Tfert),  np.column_stack((mass1,mass2))
    elif trials == 3:
        d,d1 = [],[]
        m,m1 = [],[]
        y,y1 = [],[]
        Obs1,mass1 = [],[]
        Obs2,mass2 = [],[]
        Obs3,mass3 = [],[]
        k = 0
        for eachLine in fileList:
            date,obs1,obs2,obs3 = eachLine.split(',')
            if len(date) >= 1 and k < 0.5:
                t0,t1,t2=date.split('/')
                d.append(float(t0))
                m.append(float(t1))
                y.append(float(t2))
                Obs1.append(float(obs1))
                Obs2.append(float(obs2))
                Obs3.append(float(obs3))
            elif len(date) < 1 and k < 0.5:
                k = k + 1
            elif len(date) >= 1 and k > 0.5:
                t0,t1,t2=date.split('/')
                d1.append(float(t0))
                m1.append(float(t1))
                y1.append(float(t2))
                mass1.append(float(obs1))
                mass2.append(float(obs2))
                mass3.append(float(obs3))
        TObs = [xx-1+30*(yy-1)+365*zz for xx,yy,zz in zip (d, m, y)]
        Tfert = [xx-1+30*(yy-1)+365*zz for xx,yy,zz in zip (d1, m1, y1)]
        return np.array(TObs), np.column_stack((Obs1,Obs2,Obs3)),np.array(Tfert),  np.column_stack((mass1,mass2,mass3))

    elif trials == 4:
        d,d1 = [],[]
        m,m1 = [],[]
        y,y1 = [],[]
        Obs1,mass1 = [],[]
        Obs2,mass2 = [],[]
        Obs3,mass3 = [],[]
        Obs4,mass4 = [],[]
        k = 0
        for eachLine in fileList:
            date,obs1,obs2,obs3,obs4 = eachLine.split(',')
            if len(date) >= 1 and k < 0.5:
                t0,t1,t2=date.split('/')
                d.append(float(t0))
                m.append(float(t1))
                y.append(float(t2))
                Obs1.append(float(obs1))
                Obs2.append(float(obs2))
                Obs3.append(float(obs3))
                Obs4.append(float(obs4))
            elif len(date) < 1 and k < 0.5:
                k = k + 1
            elif len(date) >= 1 and k > 0.5:
                t0,t1,t2=date.split('/')
                d1.append(float(t0))
                m1.append(float(t1))
                y1.append(float(t2))
                mass1.append(float(obs1))
                mass2.append(float(obs2))
                mass3.append(float(obs3))
                mass4.append(float(obs4))
        TObs = [xx-1+30*(yy-1)+365*zz for xx,yy,zz in zip (d, m, y)]
        Tfert = [xx-1+30*(yy-1)+365*zz for xx,yy,zz in zip (d1, m1, y1)]
        return np.array(TObs), np.column_stack((Obs1,Obs2,Obs3,Obs4)),np.array(Tfert),  np.column_stack((mass1,mass2,mass3,mass4))
        
    elif trials == 5:
        d,d1 = [],[]
        m,m1 = [],[]
        y,y1 = [],[]
        Obs1,mass1 = [],[]
        Obs2,mass2 = [],[]
        Obs3,mass3 = [],[]
        Obs4,mass4 = [],[]
        Obs5,mass5 = [],[]
        k = 0
        for eachLine in fileList:
            date,obs1,obs2,obs3,obs4,obs5 = eachLine.split(',')
            if len(date) >= 1 and k < 0.5:
                t0,t1,t2=date.split('/')
                d.append(float(t0))
                m.append(float(t1))
                y.append(float(t2))
                Obs1.append(float(obs1))
                Obs2.append(float(obs2))
                Obs3.append(float(obs3))
                Obs4.append(float(obs4))
                Obs5.append(float(obs5))
            elif len(date) < 1 and k < 0.5:
                k = k + 1
            elif len(date) >= 1 and k > 0.5:
                t0,t1,t2=date.split('/')
                d1.append(float(t0))
                m1.append(float(t1))
                y1.append(float(t2))
                mass1.append(float(obs1))
                mass2.append(float(obs2))
                mass3.append(float(obs3))
                mass4.append(float(obs4))
                mass5.append(float(obs5))
        TObs = [xx-1+30*(yy-1)+365*zz for xx,yy,zz in zip (d, m, y)]
        Tfert = [xx-1+30*(yy-1)+365*zz for xx,yy,zz in zip (d1, m1, y1)]
        return np.array(TObs), np.column_stack((Obs1,Obs2,Obs3,Obs4,Obs5)),np.array(Tfert),  np.column_stack((mass1,mass2,mass3,mass4,mass5))

    elif trials == 6:
        d,d1 = [],[]
        m,m1 = [],[]
        y,y1 = [],[]
        Obs1,mass1 = [],[]
        Obs2,mass2 = [],[]
        Obs3,mass3 = [],[]
        Obs4,mass4 = [],[]
        Obs5,mass5 = [],[]
        Obs6,mass6 = [],[]
        k = 0
        for eachLine in fileList:
            date,obs1,obs2,obs3,obs4,obs5,obs6 = eachLine.split(',')
            if len(date) >= 1 and k < 0.5:
                t0,t1,t2=date.split('/')
                d.append(float(t0))
                m.append(float(t1))
                y.append(float(t2))
                Obs1.append(float(obs1))
                Obs2.append(float(obs2))
                Obs3.append(float(obs3))
                Obs4.append(float(obs4))
                Obs5.append(float(obs5))
                Obs6.append(float(obs6))
            elif len(date) < 1 and k < 0.5:
                k = k + 1
            elif len(date) >= 1 and k > 0.5:
                t0,t1,t2=date.split('/')
                d1.append(float(t0))
                m1.append(float(t1))
                y1.append(float(t2))
                mass1.append(float(obs1))
                mass2.append(float(obs2))
                mass3.append(float(obs3))
                mass4.append(float(obs4))
                mass5.append(float(obs5))
                mass6.append(float(obs6))
        TObs = [xx-1+30*(yy-1)+365*zz for xx,yy,zz in zip (d, m, y)]
        Tfert = [xx-1+30*(yy-1)+365*zz for xx,yy,zz in zip (d1, m1, y1)]
        return np.array(TObs), np.column_stack((Obs1,Obs2,Obs3,Obs4,Obs5,Obs6)),np.array(Tfert),  np.column_stack((mass1,mass2,mass3,mass4,mass5,mass6))
    else:
        return 'No Data'


def PKXl():
    fileName = askopenfilename( filetypes = [("Excel files", "*.xlsx"),("All files", "*.*")])
    f = load_workbook(fileName)
    sheet=f.active
    trials = int((sheet.max_column - 2)/2)
    TotalR = sheet.max_row
    if trials == 1:
        Tobs, Obs, Tfert, mass = [],[],[],[]
        for i in range(2,TotalR+1):
            try:
                Tobs.append(sheet.cell(row=i, column=1).value.day+30*(sheet.cell(row=i, column=1).value.month-1)+sheet.cell(row=i, column=1).value.year*365)
            except(AttributeError):
                pass
            try:
                Obs.append(sheet.cell(row=i, column=2).value)
            except(AttributeError):
                pass
            try:
                Tfert.append(sheet.cell(row=i, column=3).value.day+30*(sheet.cell(row=i, column=3).value.month-1)+sheet.cell(row=i, column=3).value.year*365)
            except(AttributeError):
                pass
            try:
                mass.append(sheet.cell(row=i, column=4).value)
            except(AttributeError):
                pass
        return np.array(Tobs),np.array(Obs), np.array(Tfert), np.array(mass)
    elif trials == 2:
        Tobs, Obs1, Obs2, Tfert, mass1, mass2 = [],[],[],[],[],[]
        for i in range(2,TotalR+1):
            try:
                Tobs.append(sheet.cell(row=i, column=1).value.day+30*(sheet.cell(row=i, column=1).value.month-1)+sheet.cell(row=i, column=1).value.year*365)
            except(AttributeError):
                pass
            try:
                Obs1.append(sheet.cell(row=i, column=2).value)
            except(AttributeError):
                pass
            try:
                Obs2.append(sheet.cell(row=i, column=3).value)
            except(AttributeError):
                pass
            try:
                Tfert.append(sheet.cell(row=i, column=4).value.day+30*(sheet.cell(row=i, column=4).value.month-1)+sheet.cell(row=i, column=4).value.year*365)
            except(AttributeError):
                pass
            try:
                mass1.append(sheet.cell(row=i, column=5).value)
            except(AttributeError):
                pass
            try:
                mass2.append(sheet.cell(row=i, column=6).value)
            except(AttributeError):
                pass
        return np.array(Tobs), np.column_stack((Obs1,Obs2)),np.array(Tfert),  np.column_stack((mass1,mass2))

    elif trials == 3:
        Tobs, Obs1, Obs2, Obs3, Tfert, mass1, mass2, mass3 = [],[],[],[],[],[],[],[]
        for i in range(2,TotalR+1):
            try:
                Tobs.append(sheet.cell(row=i, column=1).value.day+30*(sheet.cell(row=i, column=1).value.month-1)+sheet.cell(row=i, column=1).value.year*365)
            except(AttributeError):
                pass
            try:
                Obs1.append(sheet.cell(row=i, column=2).value)
            except(AttributeError):
                pass
            try:
                Obs2.append(sheet.cell(row=i, column=3).value)
            except(AttributeError):
                pass
            try:
                Obs3.append(sheet.cell(row=i, column=4).value)
            except(AttributeError):
                pass
            try:
                Tfert.append(sheet.cell(row=i, column=5).value.day+30*(sheet.cell(row=i, column=5).value.month-1)+sheet.cell(row=i, column=5).value.year*365)
            except(AttributeError):
                pass
            try:
                mass1.append(sheet.cell(row=i, column=6).value)
            except(AttributeError):
                pass
            try:
                mass2.append(sheet.cell(row=i, column=7).value)
            except(AttributeError):
                pass
            try:
                mass3.append(sheet.cell(row=i, column=8).value)
            except(AttributeError):
                pass
        return np.array(Tobs), np.column_stack((Obs1,Obs2,Obs3)),np.array(Tfert),  np.column_stack((mass1,mass2,mass3))

    elif trials == 4:
        Tobs, Obs1, Obs2, Obs3, Obs4, Tfert, mass1, mass2, mass3, mass4 = [],[],[],[],[],[],[],[],[],[]
        for i in range(2,TotalR+1):
            try:
                Tobs.append(sheet.cell(row=i, column=1).value.day+30*(sheet.cell(row=i, column=1).value.month-1)+sheet.cell(row=i, column=1).value.year*365)
            except(AttributeError):
                pass
            try:
                Obs1.append(sheet.cell(row=i, column=2).value)
            except(AttributeError):
                pass
            try:
                Obs2.append(sheet.cell(row=i, column=3).value)
            except(AttributeError):
                pass
            try:
                Obs3.append(sheet.cell(row=i, column=4).value)
            except(AttributeError):
                pass
            try:
                Obs4.append(sheet.cell(row=i, column=5).value)
            except(AttributeError):
                pass
            try:
                Tfert.append(sheet.cell(row=i, column=6).value.day+30*(sheet.cell(row=i, column=6).value.month-1)+sheet.cell(row=i, column=6).value.year*365)
            except(AttributeError):
                pass
            try:
                mass1.append(sheet.cell(row=i, column=7).value)
            except(AttributeError):
                pass
            try:
                mass2.append(sheet.cell(row=i, column=8).value)
            except(AttributeError):
                pass
            try:
                mass3.append(sheet.cell(row=i, column=9).value)
            except(AttributeError):
                pass
            try:
                mass4.append(sheet.cell(row=i, column=10).value)
            except(AttributeError):
                pass
        return np.array(Tobs), np.column_stack((Obs1,Obs2,Obs3,Obs4)),np.array(Tfert),  np.column_stack((mass1,mass2,mass3,mass4))
    elif trials == 5:
        Tobs, Obs1, Obs2, Obs3, Obs4, Obs5, Tfert, mass1, mass2, mass3, mass4, mass5 = [],[],[],[],[],[],[],[],[],[],[],[]
        for i in range(2,TotalR+1):
            try:
                Tobs.append(sheet.cell(row=i, column=1).value.day+30*(sheet.cell(row=i, column=1).value.month-1)+sheet.cell(row=i, column=1).value.year*365)
            except(AttributeError):
                pass
            try:
                Obs1.append(sheet.cell(row=i, column=2).value)
            except(AttributeError):
                pass
            try:
                Obs2.append(sheet.cell(row=i, column=3).value)
            except(AttributeError):
                pass
            try:
                Obs3.append(sheet.cell(row=i, column=4).value)
            except(AttributeError):
                pass
            try:
                Obs4.append(sheet.cell(row=i, column=5).value)
            except(AttributeError):
                pass
            try:
                Obs5.append(sheet.cell(row=i, column=6).value)
            except(AttributeError):
                pass
            try:
                Tfert.append(sheet.cell(row=i, column=7).value.day+30*(sheet.cell(row=i, column=7).value.month-1)+sheet.cell(row=i, column=7).value.year*365)
            except(AttributeError):
                pass
            try:
                mass1.append(sheet.cell(row=i, column=8).value)
            except(AttributeError):
                pass
            try:
                mass2.append(sheet.cell(row=i, column=9).value)
            except(AttributeError):
                pass
            try:
                mass3.append(sheet.cell(row=i, column=10).value)
            except(AttributeError):
                pass
            try:
                mass4.append(sheet.cell(row=i, column=11).value)
            except(AttributeError):
                pass
            try:
                mass5.append(sheet.cell(row=i, column=12).value)
            except(AttributeError):
                pass
        return np.array(Tobs), np.column_stack((Obs1,Obs2,Obs3,Obs4,Obs5)),np.array(Tfert),  np.column_stack((mass1,mass2,mass3,mass4,mass5))

    elif trials == 6:
        Tobs, Obs1, Obs2, Obs3, Obs4, Obs5, Obs6, Tfert, mass1, mass2, mass3, mass4, mass5, mass6= [],[],[],[],[],[],[],[],[],[],[],[],[],[]
        for i in range(2,TotalR+1):
            try:
                Tobs.append(sheet.cell(row=i, column=1).value.day+30*(sheet.cell(row=i, column=1).value.month-1)+sheet.cell(row=i, column=1).value.year*365)
            except(AttributeError):
                pass
            try:
                Obs1.append(sheet.cell(row=i, column=2).value)
            except(AttributeError):
                pass
            try:
                Obs2.append(sheet.cell(row=i, column=3).value)
            except(AttributeError):
                pass
            try:
                Obs3.append(sheet.cell(row=i, column=4).value)
            except(AttributeError):
                pass
            try:
                Obs4.append(sheet.cell(row=i, column=5).value)
            except(AttributeError):
                pass
            try:
                Obs5.append(sheet.cell(row=i, column=6).value)
            except(AttributeError):
                pass
            try:
                Obs6.append(sheet.cell(row=i, column=7).value)
            except(AttributeError):
                pass
            try:
                Tfert.append(sheet.cell(row=i, column=8).value.day+30*(sheet.cell(row=i, column=8).value.month-1)+sheet.cell(row=i, column=8).value.year*365)
            except(AttributeError):
                pass
            try:
                mass1.append(sheet.cell(row=i, column=9).value)
            except(AttributeError):
                pass
            try:
                mass2.append(sheet.cell(row=i, column=10).value)
            except(AttributeError):
                pass
            try:
                mass3.append(sheet.cell(row=i, column=11).value)
            except(AttributeError):
                pass
            try:
                mass4.append(sheet.cell(row=i, column=12).value)
            except(AttributeError):
                pass
            try:
                mass5.append(sheet.cell(row=i, column=13).value)
            except(AttributeError):
                pass
            try:
                mass6.append(sheet.cell(row=i, column=14).value)
            except(AttributeError):
                pass
        return np.array(Tobs), np.column_stack((Obs1,Obs2,Obs3,Obs4,Obs5,Obs6)),np.array(Tfert),  np.column_stack((mass1,mass2,mass3,mass4,mass5,mass6))
    else:
        return 'No Data'
